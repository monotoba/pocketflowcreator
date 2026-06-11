"""Generate fully standalone, self-contained Python scripts that run a flow without pocketflow."""

from __future__ import annotations

import re
from pathlib import Path
from typing import TYPE_CHECKING

from pocketflow_creator.model.graph_model import GraphModel

if TYPE_CHECKING:
    from pocketflow_creator.model.provider_profile import ProjectProviders, ProviderProfile


class StandaloneGenerator:
    """Generate a complete, stdlib-only self-contained Python script."""

    # External dependency registry: maps node_type_id to dependency info
    # Format: {node_type: {"lib": lib_name, "env_vars": [var1, var2], "program": program_name, "link": url}}
    _DEPENDENCY_REGISTRY: dict[str, dict[str, object]] = {
        "shell_command_node": {
            "program": "bash/sh/zsh/PowerShell/cmd",
            "description": "Shell command execution",
            "link": "https://www.gnu.org/software/bash/",
        },
        "tty_serial_node": {
            "lib": "pyserial",
            "program": "Serial port",
            "description": "Serial port communication",
            "link": "https://pypi.org/project/pyserial/",
        },
        "speech_to_text_node": {
            "lib": "SpeechRecognition",
            "description": "Speech recognition",
            "link": "https://pypi.org/project/SpeechRecognition/",
        },
        "text_to_speech_node": {
            "lib": "pyttsx3",
            "description": "Text-to-speech synthesis",
            "link": "https://pypi.org/project/pyttsx3/",
        },
        "pdf_extract_node": {
            "lib": "PyPDF2",
            "description": "PDF text extraction",
            "link": "https://pypi.org/project/PyPDF2/",
        },
        "spreadsheet_node": {
            "lib": "openpyxl",
            "description": "Excel file support",
            "link": "https://pypi.org/project/openpyxl/",
        },
        "web_scrape_node": {
            "lib": "beautifulsoup4",
            "description": "Web scraping",
            "link": "https://pypi.org/project/beautifulsoup4/",
        },
        "websocket_node": {
            "lib": "websockets",
            "description": "WebSocket communication",
            "link": "https://pypi.org/project/websockets/",
        },
        "web_search_node": {
            "env_vars": ["SEARCH_API_KEY"],
            "description": "Web search API",
        },
        "email_send_node": {
            "env_vars": ["EMAIL_ADDRESS", "EMAIL_PASSWORD", "SMTP_SERVER"],
            "description": "Email sending",
        },
        "email_read_node": {
            "env_vars": ["EMAIL_ADDRESS", "EMAIL_PASSWORD", "IMAP_SERVER"],
            "description": "Email reading",
        },
        "calendar_read_node": {
            "env_vars": ["GOOGLE_CALENDAR_ID"],
            "description": "Google Calendar integration",
            "link": "https://developers.google.com/calendar/api",
        },
        "calendar_write_node": {
            "env_vars": ["GOOGLE_CALENDAR_ID"],
            "description": "Google Calendar integration",
            "link": "https://developers.google.com/calendar/api",
        },
        "socket_node": {
            "env_vars": ["SOCKET_HOST", "SOCKET_PORT"],
            "description": "Socket communication",
        },
        "notification_node": {
            "env_vars": ["SLACK_WEBHOOK", "DISCORD_WEBHOOK"],
            "description": "Notification services",
        },
        "secret_node": {
            "env_vars": ["SECRET_SOURCE"],
            "description": "Secret management",
        },
        "mcp_tool_node": {
            "env_vars": ["MCP_SERVER_URL"],
            "description": "MCP server integration",
        },
        "python_tool_node": {
            "description": "Python module import",
        },
    }

    # Provider source code embedded as strings (must stay in sync with providers.py)
    _PROVIDER_SOURCES: dict[str, str] = {
        "ollama": '''\
class OllamaProvider:
    """Ollama native API — /api/generate endpoint."""

    def __init__(self, base_url: str = "http://localhost:11434", default_model: str = "qwen2.5-coder:14b", timeout: int = 120):
        self.base_url = base_url
        self.default_model = default_model
        self.timeout = timeout

    def complete(self, prompt: str, *, model: str | None = None) -> str:
        import sys
        url = f"{self.base_url.rstrip('/')}/api/generate"
        model_name = model or self.default_model
        payload = json.dumps({"model": model_name, "prompt": prompt, "stream": False}).encode()
        print(f"[OllamaProvider] URL: {url}", file=sys.stderr)
        print(f"[OllamaProvider] Request: {payload.decode()}", file=sys.stderr)
        req = urllib.request.Request(url, data=payload, headers={"Content-Type": "application/json"}, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=self.timeout) as resp:
                body = json.loads(resp.read())
                return str(body.get("response", ""))
        except urllib.error.HTTPError as exc:
            print(f"[OllamaProvider] HTTPError: {exc.code} {exc.reason}", file=sys.stderr)
            detail = f"{exc.code} {exc.reason}"
            try:
                resp_body = exc.read().decode("utf-8", errors="replace")
                print(f"[OllamaProvider] Response body: {resp_body[:500]}", file=sys.stderr)
                if resp_body:
                    try:
                        error_data = json.loads(resp_body)
                        detail = error_data.get("error", resp_body[:200])
                    except json.JSONDecodeError:
                        detail = resp_body[:200]
            except Exception as e:
                print(f"[OllamaProvider] Error reading response: {e}", file=sys.stderr)
                pass
            detail = detail or f"{exc.code} {exc.reason}"
            if "not found" in detail.lower() or "unknown" in detail.lower():
                detail = f"{detail}. Make sure '{model_name}' is installed in Ollama (ollama pull {model_name})"
            raise RuntimeError(f"Ollama request failed: {detail}") from exc
        except urllib.error.URLError as exc:
            reason = str(exc.reason) if exc.reason else str(exc)
            print(f"[OllamaProvider] URLError: {reason}", file=sys.stderr)
            raise RuntimeError(f"Ollama request failed: {reason}") from exc
        except (json.JSONDecodeError, KeyError, ValueError) as exc:
            print(f"[OllamaProvider] Response parsing error: {exc}", file=sys.stderr)
            raise RuntimeError(f"Ollama returned unexpected response: {exc}") from exc
''',
        "openai": '''\
class OpenAIProvider:
    """OpenAI chat completions — also works with any OpenAI-compatible endpoint."""

    def __init__(self, api_key: str, base_url: str = "https://api.openai.com/v1", default_model: str = "gpt-4o-mini", timeout: int = 120):
        self.api_key = api_key
        self.base_url = base_url
        self.default_model = default_model
        self.timeout = timeout

    def complete(self, prompt: str, *, model: str | None = None) -> str:
        url = f"{self.base_url.rstrip('/')}/chat/completions"
        payload = json.dumps({
            "model": model or self.default_model,
            "messages": [{"role": "user", "content": prompt}],
        }).encode()
        headers = {"Content-Type": "application/json"}
        if self.api_key:
            headers["Authorization"] = f"Bearer {self.api_key}"
        req = urllib.request.Request(url, data=payload, headers=headers, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=self.timeout) as resp:
                body = json.loads(resp.read())
                return str(body["choices"][0]["message"]["content"])
        except urllib.error.HTTPError as exc:
            try:
                detail = json.loads(exc.read()).get("error", {}).get("message", str(exc))
            except Exception:
                detail = str(exc)
            raise RuntimeError(f"OpenAI request failed: {detail}") from exc
        except urllib.error.URLError as exc:
            reason = str(exc.reason) if exc.reason else str(exc)
            raise RuntimeError(f"OpenAI request failed: {reason}") from exc
        except (json.JSONDecodeError, KeyError, IndexError, ValueError) as exc:
            raise RuntimeError(f"OpenAI returned unexpected response: {exc}") from exc
''',
        "anthropic": '''\
class AnthropicProvider:
    """Anthropic Messages API."""

    def __init__(self, api_key: str, default_model: str = "claude-haiku-4-5", timeout: int = 120):
        self.api_key = api_key
        self.default_model = default_model
        self.timeout = timeout

    def complete(self, prompt: str, *, model: str | None = None) -> str:
        url = "https://api.anthropic.com/v1/messages"
        payload = json.dumps({
            "model": model or self.default_model,
            "max_tokens": 4096,
            "messages": [{"role": "user", "content": prompt}],
        }).encode()
        req = urllib.request.Request(
            url,
            data=payload,
            headers={
                "Content-Type": "application/json",
                "x-api-key": self.api_key,
                "anthropic-version": "2023-06-01",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=self.timeout) as resp:
                body = json.loads(resp.read())
                return str(body["content"][0]["text"])
        except urllib.error.HTTPError as exc:
            try:
                detail = json.loads(exc.read()).get("error", {}).get("message", str(exc))
            except Exception:
                detail = str(exc)
            raise RuntimeError(f"Anthropic request failed: {detail}") from exc
        except urllib.error.URLError as exc:
            reason = str(exc.reason) if exc.reason else str(exc)
            raise RuntimeError(f"Anthropic request failed: {reason}") from exc
        except (json.JSONDecodeError, KeyError, IndexError) as exc:
            raise RuntimeError(f"Anthropic returned unexpected response: {exc}") from exc
''',
        "gemini": '''\
class GeminiProvider:
    """Google Gemini generateContent API."""

    def __init__(self, api_key: str, default_model: str = "gemini-2.0-flash", timeout: int = 120):
        self.api_key = api_key
        self.default_model = default_model
        self.timeout = timeout

    def complete(self, prompt: str, *, model: str | None = None) -> str:
        mdl = model or self.default_model
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{mdl}:generateContent?key={self.api_key}"
        payload = json.dumps({"contents": [{"parts": [{"text": prompt}]}]}).encode()
        req = urllib.request.Request(
            url,
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=self.timeout) as resp:
                body = json.loads(resp.read())
                return str(body["candidates"][0]["content"]["parts"][0]["text"])
        except urllib.error.HTTPError as exc:
            try:
                detail = json.loads(exc.read()).get("error", {}).get("message", str(exc))
            except Exception:
                detail = str(exc)
            raise RuntimeError(f"Gemini request failed: {detail}") from exc
        except urllib.error.URLError as exc:
            raise RuntimeError(f"Gemini request failed: {exc}") from exc
        except (json.JSONDecodeError, KeyError, IndexError) as exc:
            raise RuntimeError(f"Gemini returned unexpected response: {exc}") from exc
''',
        "deepseek": '''\
class DeepSeekProvider:
    """DeepSeek API (OpenAI-compatible)."""

    def __init__(self, api_key: str, base_url: str = "https://api.deepseek.com/v1", default_model: str = "deepseek-chat", timeout: int = 120):
        self.api_key = api_key
        self.base_url = base_url
        self.default_model = default_model
        self.timeout = timeout

    def complete(self, prompt: str, *, model: str | None = None) -> str:
        url = f"{self.base_url.rstrip('/')}/chat/completions"
        payload = json.dumps({
            "model": model or self.default_model,
            "messages": [{"role": "user", "content": prompt}],
        }).encode()
        headers = {"Content-Type": "application/json"}
        if self.api_key:
            headers["Authorization"] = f"Bearer {self.api_key}"
        req = urllib.request.Request(url, data=payload, headers=headers, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=self.timeout) as resp:
                body = json.loads(resp.read())
                return str(body["choices"][0]["message"]["content"])
        except urllib.error.HTTPError as exc:
            try:
                detail = json.loads(exc.read()).get("error", {}).get("message", str(exc))
            except Exception:
                detail = str(exc)
            raise RuntimeError(f"DeepSeek request failed: {detail}") from exc
        except urllib.error.URLError as exc:
            raise RuntimeError(f"DeepSeek request failed: {exc}") from exc
        except (json.JSONDecodeError, KeyError, IndexError, ValueError) as exc:
            raise RuntimeError(f"DeepSeek returned unexpected response: {exc}") from exc
''',
        "failover": '''\
class FailoverProvider:
    """Composite provider that retries across multiple providers on failure."""

    def __init__(self, entries, cooldowns=None):
        self.entries = sorted(entries, key=lambda e: e.get("priority", 999))
        self._cooldowns = cooldowns or {}
        self._time = __import__("time")

    def complete(self, prompt: str, *, model: str | None = None) -> str:
        """Try each provider in priority order with per-error-type retries."""
        now = self._time.time()
        last_error = None

        for entry in self.entries:
            provider_key = str(id(entry["provider"]))
            if provider_key in self._cooldowns and now < self._cooldowns[provider_key]:
                continue

            provider = entry["provider"]
            provider_model = entry.get("model") or model
            max_retries = {
                "timeout": entry.get("timeout_retries", 3),
                "network": entry.get("network_retries", 3),
                "ratelimit": entry.get("ratelimit_retries", 2),
                "expired": entry.get("expired_retries", 1),
                "unknown": entry.get("unknown_retries", 1),
            }

            for attempt in range(max(max_retries.values()) + 1):
                try:
                    return provider.complete(prompt, model=provider_model)
                except Exception as exc:
                    last_error = exc
                    error_type = "unknown"
                    if "timeout" in str(type(exc).__name__).lower():
                        error_type = "timeout"
                    elif "network" in str(type(exc).__name__).lower() or "url" in str(type(exc).__name__).lower():
                        error_type = "network"
                    elif "429" in str(exc):
                        error_type = "ratelimit"
                    elif "401" in str(exc) or "403" in str(exc) or "expired" in str(exc).lower():
                        error_type = "expired"

                    if attempt < max_retries[error_type]:
                        self._time.sleep(entry.get("retry_delay", 2.0))
                    else:
                        break

        raise RuntimeError(f"All failover providers exhausted. Last error: {last_error}")
''',
    }

    def collect_dependencies(self, graph: GraphModel) -> dict[str, str]:
        """Collect all pip-installable dependencies required by this graph.

        Returns dict of {package_name: package_spec} (e.g., {"PyPDF2": "PyPDF2>=3.0"})
        """
        dependencies: dict[str, str] = {}

        for node in graph.nodes:
            node_type = node.type_id
            if node_type in self._DEPENDENCY_REGISTRY:
                dep_info = self._DEPENDENCY_REGISTRY[node_type]
                if "lib" in dep_info:
                    lib_name = str(dep_info["lib"])
                    # Map pip package name to common version specs
                    version_specs = {
                        "pyserial": "pyserial>=3.5",
                        "beautifulsoup4": "beautifulsoup4>=4.11",
                        "websockets": "websockets>=11.0",
                        "PyPDF2": "PyPDF2>=3.0",
                        "openpyxl": "openpyxl>=3.10",
                        "pyttsx3": "pyttsx3>=2.90",
                        "SpeechRecognition": "SpeechRecognition>=3.10",
                        "sounddevice": "sounddevice>=0.4.5",
                        "soundfile": "soundfile>=0.12.1",
                        "opencv-python": "opencv-python>=4.8.0",
                    }
                    dependencies[lib_name] = version_specs.get(lib_name, lib_name)

        return dependencies

    def generate(
        self,
        graph: GraphModel,
        project_providers: ProjectProviders,
        project_name: str = "",
        project_root: Path | None = None,
    ) -> str:
        """Generate a complete standalone Python script for the graph."""
        # Collect which providers are actually used
        used_profiles = self._collect_used_profiles(graph, project_providers)
        used_types = {p.type for p in used_profiles.values()}

        # Build mappings
        profile_var_map = {p_id: self._safe_var_name(p.name) for p_id, p in used_profiles.items()}

        # Generate sections
        header = self._render_header(project_name, graph.title)
        imports = self._render_imports()
        provider_classes = self._render_provider_classes(used_types, graph)
        provider_instances = self._render_provider_instances(used_profiles, profile_var_map)
        helpers = self._render_helpers()
        graph_data = self._render_graph_data(graph, used_profiles, profile_var_map)
        node_dispatch = self._render_node_dispatch()
        run_flow = self._render_run_flow()
        main_block = self._render_main_block(graph)

        return "\n\n".join([header, imports, provider_classes, provider_instances, helpers, graph_data, node_dispatch, run_flow, main_block])

    def _collect_used_profiles(self, graph: GraphModel, project_providers: ProjectProviders) -> dict[str, ProviderProfile]:
        """Return {profile_id: ProviderProfile} for all providers referenced in the graph."""
        used_ids = set()
        for node in graph.nodes:
            pid = str(node.properties.get("provider_id", "")).strip()
            if pid:
                used_ids.add(pid)

        result = {}
        for profile in project_providers.profiles:
            if profile.id in used_ids:
                result[profile.id] = profile

        # If no profiles were explicitly referenced, add the default profile
        if not result and project_providers.default_profile:
            result[project_providers.default_profile.id] = project_providers.default_profile

        return result

    def _safe_var_name(self, name: str) -> str:
        """Convert a profile name to a safe Python variable name."""
        safe = re.sub(r"[^a-zA-Z0-9_]", "_", name.lower())
        # Remove leading digits and underscores
        safe = re.sub(r"^[0-9_]+", "", safe)
        return f"_provider_{safe}" if safe else "_provider_default"

    def check_dependencies(self, graph: GraphModel) -> dict[str, dict[str, object]]:
        """Check graph for external dependencies and return missing ones.

        Returns dict mapping node_type -> dependency info for nodes with missing deps.
        """
        missing = {}
        for node in graph.nodes:
            if node.type_id not in self._DEPENDENCY_REGISTRY:
                continue
            dep_info = self._DEPENDENCY_REGISTRY[node.type_id]

            # Check environment variables
            if "env_vars" in dep_info:
                env_vars = dep_info["env_vars"]
                if isinstance(env_vars, list):
                    missing_vars = [v for v in env_vars if not __import__("os").environ.get(v)]
                    if missing_vars:
                        missing[node.type_id] = {**dep_info, "missing_env_vars": missing_vars}

            # Check for required libraries
            if "lib" in dep_info:
                lib_name = str(dep_info["lib"])
                if not self._try_import(lib_name):
                    missing[node.type_id] = {**dep_info, "missing_lib": lib_name}

        return missing

    @staticmethod
    def _try_import(module_name: str) -> bool:
        """Try to import a module, return True if successful."""
        # Handle package names that differ from import names
        import_map = {
            "PyPDF2": "PyPDF2",
            "beautifulsoup4": "bs4",
            "openpyxl": "openpyxl",
            "pyserial": "serial",
        }
        import_name = import_map.get(module_name, module_name)
        try:
            __import__(import_name)
            return True
        except ImportError:
            return False

    def _render_header(self, project_name: str, graph_title: str) -> str:
        """Render the file header comment."""
        lines = [
            "# ──────────────────────────────────────────────────────────────────────────────────",
            "# Auto-generated by PocketFlow Creator — do not edit",
            f"# Project: {project_name}",
            f"# Graph: {graph_title}",
            "# ──────────────────────────────────────────────────────────────────────────────────",
        ]
        return "\n".join(lines)

    def _render_imports(self) -> str:
        """Render the import block."""
        return """\
from __future__ import annotations

import base64
import copy
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path"""

    def _render_provider_classes(self, used_types: set[str], graph: GraphModel | None = None) -> str:
        """Render provider class definitions for used types."""
        lines = ["# ── Provider Classes ─────────────────────────────────────────────────────────"]

        # Determine which provider type names map to each class
        type_to_provider_class: dict[str, str] = {
            "ollama": "ollama",
            "lm_studio": "openai",
            "openai_compat": "openai",
            "deepseek": "deepseek",
            "anthropic": "anthropic",
            "gemini": "gemini",
        }

        rendered_classes = set()
        for ptype in used_types:
            provider_key = type_to_provider_class.get(ptype)
            if provider_key and provider_key not in rendered_classes:
                lines.append(self._PROVIDER_SOURCES[provider_key])
                rendered_classes.add(provider_key)

        # Check if FailoverProvider is needed
        if graph:
            has_failover = any(n.type_id == "provider_failover_node" for n in graph.nodes)
            if has_failover and "failover" not in rendered_classes:
                lines.append(self._PROVIDER_SOURCES["failover"])
                rendered_classes.add("failover")

        return "\n".join(lines)

    def _render_provider_instances(self, profiles: dict[str, ProviderProfile], profile_var_map: dict[str, str]) -> str:
        """Render provider instance declarations."""
        lines = ["# ── Provider Instances ───────────────────────────────────────────────────────"]
        lines.append("")

        if not profiles:
            return "\n".join(lines)

        # Check if any cloud provider keys are needed
        cloud_profiles = [p for p in profiles.values() if p.type in ("openai_compat", "anthropic", "gemini", "deepseek", "lm_studio")]
        if any(p.type in ("openai_compat", "anthropic", "gemini", "deepseek") for p in cloud_profiles):
            lines.append("# ── API Keys ──────────────────────────────────────────────────────────────────")
            lines.append("# Set these environment variables before running:")
            seen = set()
            for p in sorted(cloud_profiles, key=lambda x: x.type):
                if p.type not in seen:
                    if p.type == "openai_compat":
                        lines.append("#   export OPENAI_API_KEY='sk-...'")
                        seen.add("openai_compat")
                    elif p.type == "anthropic":
                        lines.append("#   export ANTHROPIC_API_KEY='sk-ant-...'")
                        seen.add("anthropic")
                    elif p.type == "gemini":
                        lines.append("#   export GEMINI_API_KEY='...'")
                        seen.add("gemini")
                    elif p.type == "deepseek":
                        lines.append("#   export DEEPSEEK_API_KEY='sk-...'")
                        seen.add("deepseek")
            lines.append("")

        for p_id, profile in profiles.items():
            var_name = profile_var_map[p_id]
            if profile.type == "ollama":
                base_url = profile.base_url or "http://localhost:11434"
                model = profile.model or "qwen2.5-coder:14b"
                lines.append(f"{var_name} = OllamaProvider(")
                lines.append(f'    base_url="{base_url}",')
                lines.append(f'    default_model="{model}",')
                lines.append(f"    timeout={profile.timeout},")
                lines.append(")")
            elif profile.type == "lm_studio":
                base_url = profile.base_url or "http://localhost:1234/v1"
                model = profile.model or "meta-llama-3.1-8b"
                lines.append(f"{var_name} = OpenAIProvider(")
                lines.append("    api_key=os.environ.get('OPENAI_API_KEY', ''),")
                lines.append(f'    base_url="{base_url}",')
                lines.append(f'    default_model="{model}",')
                lines.append(f"    timeout={profile.timeout},")
                lines.append(")")
            elif profile.type == "openai_compat":
                base_url = profile.base_url or "https://api.openai.com/v1"
                model = profile.model or "gpt-4o-mini"
                lines.append(f"{var_name} = OpenAIProvider(")
                lines.append("    api_key=os.environ.get('OPENAI_API_KEY', ''),")
                lines.append(f'    base_url="{base_url}",')
                lines.append(f'    default_model="{model}",')
                lines.append(f"    timeout={profile.timeout},")
                lines.append(")")
            elif profile.type == "anthropic":
                model = profile.model or "claude-haiku-4-5"
                lines.append(f"{var_name} = AnthropicProvider(")
                lines.append("    api_key=os.environ.get('ANTHROPIC_API_KEY', ''),")
                lines.append(f'    default_model="{model}",')
                lines.append(f"    timeout={profile.timeout},")
                lines.append(")")
            elif profile.type == "gemini":
                model = profile.model or "gemini-2.0-flash"
                lines.append(f"{var_name} = GeminiProvider(")
                lines.append("    api_key=os.environ.get('GEMINI_API_KEY', ''),")
                lines.append(f'    default_model="{model}",')
                lines.append(f"    timeout={profile.timeout},")
                lines.append(")")
            elif profile.type == "deepseek":
                base_url = profile.base_url or "https://api.deepseek.com/v1"
                model = profile.model or "deepseek-chat"
                lines.append(f"{var_name} = DeepSeekProvider(")
                lines.append("    api_key=os.environ.get('DEEPSEEK_API_KEY', ''),")
                lines.append(f'    base_url="{base_url}",')
                lines.append(f'    default_model="{model}",')
                lines.append(f"    timeout={profile.timeout},")
                lines.append(")")

        return "\n".join(lines)

    def _render_helpers(self) -> str:
        """Render helper functions (_resolve_prompt, _interpolate, etc.)."""
        return r'''\
# ── Helper Functions ──────────────────────────────────────────────────────────

def _resolve_prompt(node_props, project_root=None):
    """Return the prompt string for an LLM node, handling both string and path types."""
    prompt_type = str(node_props.get("prompt_type", "string"))
    raw = str(node_props.get("prompt_file", ""))
    if not raw:
        return ""
    if prompt_type == "path":
        if project_root is None:
            return f"(cannot read prompt file — no project root: {raw})"
        try:
            return (Path(project_root) / raw).read_text(encoding="utf-8")
        except FileNotFoundError:
            return f"(prompt file not found: {raw})"
        except Exception as exc:
            return f"(error reading {raw}: {exc})"
    return raw


def _interpolate(text, shared_store):
    """Replace shared store references in a prompt string.

    Supports:
      {key}            — replaced with str(shared_store[key])
      shared['key']    — replaced with str(shared_store[key])
    """
    def _replace_shared(m):
        key = m.group(1)
        return str(shared_store[key]) if key in shared_store else m.group(0)

    text = re.sub(r"""shared\[['"]([^'"]+)['"]\]""", _replace_shared, text)

    def _replace_brace(m):
        key = m.group(1)
        return str(shared_store[key]) if key in shared_store else m.group(0)

    text = re.sub(r"\{([^}]+)\}", _replace_brace, text)
    return text'''

    def _render_graph_data(
        self,
        graph: GraphModel,
        used_profiles: dict[str, ProviderProfile],
        profile_var_map: dict[str, str],
    ) -> str:
        """Render the graph data (_START, _NODES, _EDGES)."""
        lines = ["# ── Graph Data ────────────────────────────────────────────────────────────────"]
        lines.append("")

        # Start node
        start_node_id = graph.start_node or (graph.nodes[0].id if graph.nodes else "")
        lines.append(f'_START = "{start_node_id}"')
        lines.append("")

        # Build _NODES dict
        lines.append("_NODES = {")
        default_profile_id = list(used_profiles.keys())[0] if used_profiles else ""
        for node in graph.nodes:
            node_type = node.type_id
            node_title = node.title.replace('"', '\\"')
            node_props_repr = repr(node.properties)

            # Determine which provider to use for this node
            provider_ref = "None"
            if "llm" in node_type.lower() or node_type in ("classifier_node", "judge_node", "agent_node"):
                provider_id = str(node.properties.get("provider_id", "")).strip() or default_profile_id
                if provider_id in profile_var_map:
                    provider_ref = profile_var_map[provider_id]

            safe_id = node.id.replace('"', '\\"')
            lines.append(f'    "{safe_id}": {{')
            lines.append(f'        "type": "{node_type}",')
            lines.append(f'        "title": "{node_title}",')
            lines.append(f'        "props": {node_props_repr},')
            lines.append(f'        "provider": {provider_ref},')
            lines.append("    },")
        lines.append("}")
        lines.append("")

        # Build _EDGES list
        lines.append("_EDGES = [")
        for edge in graph.edges:
            from_id = edge.from_node.replace('"', '\\"')
            to_id = edge.to_node.replace('"', '\\"')
            action = edge.action.replace('"', '\\"')
            lines.append(f'    {{"from": "{from_id}", "action": "{action}", "to": "{to_id}"}},')
        lines.append("]")

        return "\n".join(lines)

    def _render_node_dispatch(self) -> str:
        """Render the _run_node() dispatcher function."""
        return '''\
# ── Node Dispatch ────────────────────────────────────────────────────────────

def _run_node(node_id, node, shared, outgoing_actions):
    """Execute a single node; return the chosen next action."""
    node_type = node["type"]
    props = node["props"]
    provider = node["provider"]
    chosen_action = list(outgoing_actions)[0] if outgoing_actions else "default"

    try:
        if node_type in ("start_node", "stop_node", "basic_node"):
            pass  # no-op

        elif "llm" in node_type.lower():
            if provider is None:
                raise RuntimeError(f"No provider configured for {node_type} '{node['title']}'")
            prompt = _interpolate(_resolve_prompt(props), shared)
            if not prompt:
                prompt = f"[{node['title']}]"
            response = provider.complete(prompt, model=props.get("model"))
            output_key = str(props.get("output_key", f"{node_id}_response"))
            shared[output_key] = response

        elif node_type == "classifier_node":
            if provider is None:
                raise RuntimeError(f"No provider configured for classifier '{node['title']}'")
            input_key = str(props.get("input_key", "input"))
            content = str(shared.get(input_key, ""))
            categories = str(props.get("categories", ""))
            resolved = _resolve_prompt(props)
            prompt = _interpolate(
                resolved or f"Classify as exactly one of [{categories}].\\nReply with only the category, nothing else.\\n\\nText: {content}",
                shared,
            )
            response = provider.complete(prompt, model=props.get("model"))
            label = response.strip().lower()
            if label not in outgoing_actions:
                label = next((act for act in outgoing_actions if act in label or label in act), next(iter(outgoing_actions), "default"))
            chosen_action = label
            shared[f"{node_id}_label"] = label

        elif node_type == "judge_node":
            if provider is None:
                raise RuntimeError(f"No provider configured for judge '{node['title']}'")
            input_key = str(props.get("input_key", "content"))
            content = str(shared.get(input_key, ""))
            criteria = str(props.get("criteria", ""))
            prompt = _interpolate(
                _resolve_prompt(props) or f"Evaluate: Criteria: {criteria}\\n\\nContent: {content}\\n\\nReply: 'pass' or 'fail'",
                shared,
            )
            response = provider.complete(prompt, model=props.get("model"))
            verdict = response.strip().lower()
            if "pass" in verdict and "pass" in outgoing_actions:
                chosen_action = "pass"
            elif "fail" in verdict and "fail" in outgoing_actions:
                chosen_action = "fail"
            shared[f"{node_id}_verdict"] = verdict

        elif node_type == "agent_node":
            if provider is None:
                raise RuntimeError(f"No provider configured for agent '{node['title']}'")
            input_key = str(props.get("input_key", "task"))
            task = str(shared.get(input_key, ""))
            prompt = _interpolate(_resolve_prompt(props) or f"Complete: {task}", shared)
            response = provider.complete(prompt, model=props.get("model"))
            output_key = str(props.get("output_key", "result"))
            shared[output_key] = response
            resp_lower = response.strip().lower()
            done_words = ("done", "complete", "finished", "answer")
            if "done" in outgoing_actions and any(w in resp_lower for w in done_words):
                chosen_action = "done"
            elif "continue" in outgoing_actions:
                chosen_action = "continue"

        elif node_type == "human_input_node":
            prompt = str(props.get("prompt", "Enter input or press Enter to skip"))
            output_key = str(props.get("output_key", "input"))
            sys.stdout.write(f"\\n[{node['title']}] {prompt}\\n> ")
            sys.stdout.flush()
            try:
                user_input = sys.stdin.readline().rstrip("\\n")
                shared[output_key] = user_input
                chosen_action = "saved" if user_input else "cancelled"
            except (EOFError, KeyboardInterrupt):
                shared[f"{output_key}_error"] = "Input cancelled"
                chosen_action = "cancelled"

        elif node_type == "human_review_node":
            input_key = str(props.get("input_key", "content"))
            output_key = str(props.get("output_key", "feedback"))
            instructions = str(props.get("instructions", "Approve or reject?"))
            content = shared.get(input_key, "")
            sys.stdout.write(f"\\n[{node['title']}] {instructions}\\n")
            sys.stdout.write(f"Content:\\n{content}\\n")
            sys.stdout.write("Approve? [y/n]: ")
            sys.stdout.flush()
            try:
                verdict = sys.stdin.readline().strip().lower()
                shared[output_key] = verdict
                chosen_action = "approved" if verdict.startswith("y") else "rejected"
            except (EOFError, KeyboardInterrupt):
                shared[f"{output_key}_error"] = "Review cancelled"
                chosen_action = "rejected"

        elif node_type == "file_reader_node":
            file_path = str(props.get("file_path", ""))
            if not file_path:
                raise RuntimeError(f"file_reader_node '{node['title']}' has no file_path")
            try:
                content = Path(file_path).read_text(encoding=str(props.get("encoding", "utf-8")))
                output_key = str(props.get("output_key", "content"))
                shared[output_key] = content
            except Exception as e:
                raise RuntimeError(f"Failed to read {file_path}: {e}") from e

        elif node_type == "file_writer_node":
            file_path = str(props.get("file_path", ""))
            if not file_path:
                raise RuntimeError(f"file_writer_node '{node['title']}' has no file_path")
            input_key = str(props.get("input_key", "content"))
            content = str(shared.get(input_key, ""))
            mode = "a" if str(props.get("mode", "write")).lower() == "append" else "w"
            try:
                Path(file_path).write_text(content, mode=mode, encoding=str(props.get("encoding", "utf-8")))
            except Exception as e:
                raise RuntimeError(f"Failed to write {file_path}: {e}") from e

        elif node_type == "router_node":
            # Router returns the first available action; routing is defined by graph edges
            pass

        elif node_type == "map_node":
            items_key = str(props.get("items_key", "items"))
            items = shared.get(items_key, [])
            if not isinstance(items, list):
                items = [items]
            template = str(props.get("item_template", ""))
            output_key = str(props.get("output_key", "mapped"))
            results = []
            for item in items:
                if template and provider:
                    resp = provider.complete(template.format(item=item), model=props.get("model"))
                    results.append(resp)
                else:
                    results.append(item)
            shared[output_key] = results

        elif node_type == "reduce_node":
            items_key = str(props.get("items_key", "items"))
            items = shared.get(items_key, [])
            if not isinstance(items, list):
                items = [items]
            output_key = str(props.get("output_key", "result"))
            accumulator = props.get("initial_value", "")
            template = str(props.get("accumulator_template", ""))
            for item in items:
                if template and provider:
                    prompt = template.format(accumulator=accumulator, item=item)
                    accumulator = provider.complete(prompt, model=props.get("model"))
                else:
                    accumulator = item
            shared[output_key] = accumulator

        elif node_type == "transform_node":
            input_key = str(props.get("input_key", "input"))
            output_key = str(props.get("output_key", "output"))
            data = shared.get(input_key, {})
            transform_spec = props.get("transform_spec", {})
            if isinstance(transform_spec, dict):
                result = {}
                for out_field, in_path in transform_spec.items():
                    parts = str(in_path).split(".")
                    val = data
                    for part in parts:
                        val = val.get(part) if isinstance(val, dict) else None
                    if val is not None:
                        result[out_field] = val
                shared[output_key] = result
            else:
                shared[output_key] = data

        elif node_type == "merge_node":
            input_keys = props.get("input_keys", [])
            if not isinstance(input_keys, list):
                input_keys = [input_keys]
            output_key = str(props.get("output_key", "merged"))
            merge_strategy = str(props.get("merge_strategy", "dict"))
            if merge_strategy == "dict":
                result = {}
                for key in input_keys:
                    if key in shared:
                        result[key] = shared[key]
                shared[output_key] = result
            elif merge_strategy == "array":
                result = []
                for key in input_keys:
                    if key in shared:
                        val = shared[key]
                        result.extend(val if isinstance(val, list) else [val])
                shared[output_key] = result
            else:
                shared[output_key] = {k: shared.get(k) for k in input_keys}

        elif node_type == "condition_node":
            condition_expr = str(props.get("condition", "True"))
            try:
                result = eval(condition_expr, {"shared": shared})
                chosen_action = "true" if result else "false"
            except Exception:
                chosen_action = "false"

        elif node_type == "loop_counter_node":
            counter_key = str(props.get("counter_key", "loop_counter"))
            max_iterations = int(props.get("max_iterations", 10))
            current = int(shared.get(counter_key, 0))
            current += 1
            shared[counter_key] = current
            output_key = str(props.get("output_key", "iteration"))
            shared[output_key] = current
            chosen_action = "continue" if current < max_iterations else "done"

        elif node_type == "api_call_node":
            endpoint = str(props.get("endpoint", ""))
            method = str(props.get("method", "GET")).upper()
            payload_key = str(props.get("payload_key", "payload"))
            output_key = str(props.get("output_key", "response"))
            if not endpoint:
                raise RuntimeError("api_call_node requires 'endpoint' property")
            try:
                payload = shared.get(payload_key, {})
                if method == "GET":
                    with urllib.request.urlopen(endpoint) as resp:
                        shared[output_key] = json.loads(resp.read())
                else:
                    data = json.dumps(payload).encode() if isinstance(payload, dict) else str(payload).encode()
                    req = urllib.request.Request(endpoint, data=data, method=method, headers={"Content-Type": "application/json"})
                    with urllib.request.urlopen(req) as resp:
                        shared[output_key] = json.loads(resp.read())
            except Exception as e:
                raise RuntimeError(f"API call to {endpoint} failed: {e}") from e

        elif node_type == "code_exec_node":
            code = str(props.get("code", ""))
            output_key = str(props.get("output_key", "result"))
            if not code:
                raise RuntimeError("code_exec_node requires 'code' property")
            try:
                import io
                import contextlib
                output = io.StringIO()
                with contextlib.redirect_stdout(output):
                    exec_globals = {"shared": shared, "__builtins__": __builtins__}
                    exec(code, exec_globals)
                shared[output_key] = output.getvalue()
            except Exception as e:
                raise RuntimeError(f"Code execution failed: {e}") from e

        elif node_type == "json_parse_node":
            input_key = str(props.get("input_key", "input"))
            output_key = str(props.get("output_key", "parsed"))
            json_str = str(shared.get(input_key, "{}"))
            try:
                parsed = json.loads(json_str)
                shared[output_key] = parsed
                chosen_action = "valid"
            except json.JSONDecodeError as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "invalid"

        elif node_type == "list_ops_node":
            input_key = str(props.get("input_key", "input"))
            output_key = str(props.get("output_key", "output"))
            operation = str(props.get("operation", "length"))
            items = shared.get(input_key, [])
            if not isinstance(items, list):
                items = [items]
            if operation == "length":
                shared[output_key] = len(items)
            elif operation == "sort":
                shared[output_key] = sorted(items)
            elif operation == "reverse":
                shared[output_key] = list(reversed(items))
            elif operation == "unique":
                shared[output_key] = list(set(items))
            else:
                shared[output_key] = items

        elif node_type == "string_ops_node":
            input_key = str(props.get("input_key", "input"))
            output_key = str(props.get("output_key", "output"))
            operation = str(props.get("operation", "upper"))
            text = str(shared.get(input_key, ""))
            if operation == "upper":
                shared[output_key] = text.upper()
            elif operation == "lower":
                shared[output_key] = text.lower()
            elif operation == "capitalize":
                shared[output_key] = text.capitalize()
            elif operation == "length":
                shared[output_key] = len(text)
            elif operation == "trim":
                shared[output_key] = text.strip()
            else:
                shared[output_key] = text

        elif node_type == "log_node":
            input_key = str(props.get("input_key", "message"))
            message = str(shared.get(input_key, ""))
            log_level = str(props.get("log_level", "info")).lower()
            if log_level == "error":
                print(f"[ERROR] {message}", file=sys.stderr)
            elif log_level == "warning":
                print(f"[WARNING] {message}", file=sys.stderr)
            else:
                print(f"[INFO] {message}")

        elif node_type == "timer_node":
            import time
            duration = float(props.get("duration", 1))
            output_key = str(props.get("output_key", "elapsed"))
            start = time.time()
            time.sleep(duration)
            elapsed = time.time() - start
            shared[output_key] = elapsed

        elif node_type == "cache_node":
            cache_key = str(props.get("cache_key", "cache"))
            input_key = str(props.get("input_key", "input"))
            output_key = str(props.get("output_key", "cached"))
            if "_cache" not in shared:
                shared["_cache"] = {}
            if cache_key in shared["_cache"]:
                shared[output_key] = shared["_cache"][cache_key]
                chosen_action = "hit"
            else:
                value = shared.get(input_key)
                shared["_cache"][cache_key] = value
                shared[output_key] = value
                chosen_action = "miss"

        elif node_type == "subflow_node":
            subflow_ref = str(props.get("subflow_ref", ""))
            output_key = str(props.get("output_key", "subflow_result"))
            if not subflow_ref:
                raise RuntimeError("subflow_node requires 'subflow_ref' property")
            shared[output_key] = {"subflow": subflow_ref, "status": "referenced"}

        elif node_type == "rag_node":
            if not provider:
                raise RuntimeError("RAG node requires a provider configured")
            vector_index_key = str(props.get("vector_index_key", "vectors"))
            query_key = str(props.get("query_key", "query"))
            context_key = str(props.get("context_key", "context"))
            output_key = str(props.get("output_key", "rag_result"))
            max_context = int(props.get("max_context_items", 3))
            query = str(shared.get(query_key, ""))
            vectors = shared.get(vector_index_key, [])
            context_items = vectors[:max_context] if isinstance(vectors, list) else []
            context = " ".join(str(item) for item in context_items)
            shared[context_key] = context
            rag_prompt = "Context: " + context + " Question: " + query
            response = provider.complete(rag_prompt, model=props.get("model"))
            shared[output_key] = response

        elif node_type == "context_compact_node":
            input_key = str(props.get("input_key", "context"))
            output_key = str(props.get("output_key", "compacted"))
            strategy = str(props.get("strategy", "truncate"))
            max_length = int(props.get("max_length", 500))
            context = str(shared.get(input_key, ""))
            if strategy == "truncate":
                shared[output_key] = context[:max_length]
            elif strategy == "summarize" and provider:
                summary_prompt = "Summarize: " + context[:max_length*2]
                shared[output_key] = provider.complete(summary_prompt, model=props.get("model"))
            elif strategy == "sliding_window":
                sentences = context.split(".")
                result = ""
                for sent in sentences:
                    if len(result) + len(sent) < max_length:
                        result += sent + "."
                    else:
                        break
                shared[output_key] = result
            else:
                shared[output_key] = context[:max_length]

        elif node_type == "conversation_history_node":
            if "_conversation" not in shared:
                shared["_conversation"] = []
            history_key = str(props.get("history_key", "conversation"))
            input_key = str(props.get("input_key", "user_input"))
            max_turns = int(props.get("max_turns", 10))
            user_message = str(shared.get(input_key, ""))
            shared["_conversation"].append({"role": "user", "content": user_message})
            if len(shared["_conversation"]) > max_turns * 2:
                shared["_conversation"] = shared["_conversation"][-(max_turns*2):]
            history_parts = []
            for msg in shared["_conversation"]:
                role = str(msg.get("role", "unknown")).upper()
                content = str(msg.get("content", ""))
                history_parts.append(role + ": " + content)
            history_text = " ".join(history_parts)
            shared[history_key] = history_text

        elif node_type == "chain_of_thought_node":
            if not provider:
                raise RuntimeError("Chain of thought requires a provider")
            input_key = str(props.get("input_key", "problem"))
            output_key = str(props.get("output_key", "reasoning"))
            problem = str(shared.get(input_key, ""))
            cot_prompt = "Think step by step:\\n" + problem + "\\n\\nReasoning:"
            response = provider.complete(cot_prompt, model=props.get("model"))
            steps = [s.strip() for s in response.split("\\n") if s.strip()]
            shared[output_key] = {"steps": steps, "final": steps[-1] if steps else ""}

        elif node_type == "majority_vote_node":
            if not provider:
                raise RuntimeError("Majority vote requires a provider")
            question_key = str(props.get("question_key", "question"))
            output_key = str(props.get("output_key", "consensus"))
            num_votes = int(props.get("num_votes", 3))
            question = str(shared.get(question_key, ""))
            votes = []
            for _ in range(num_votes):
                response = provider.complete(question, model=props.get("model"))
                votes.append(response.strip()[:100])
            from collections import Counter
            vote_counts = Counter(votes)
            consensus = vote_counts.most_common(1)[0][0] if vote_counts else ""
            shared[output_key] = {"consensus": consensus, "votes": votes}

        elif node_type == "supervisor_node":
            if not provider:
                raise RuntimeError("Supervisor requires a provider")
            task_key = str(props.get("task_key", "task"))
            output_key = str(props.get("output_key", "synthesis"))
            num_agents = int(props.get("num_agents", 3))
            task = str(shared.get(task_key, ""))
            agent_responses = []
            for agent_id in range(num_agents):
                prompt = f"[Agent {agent_id}] {task}"
                response = provider.complete(prompt, model=props.get("model"))
                agent_responses.append(response)
            perspectives_text = "\\n".join(
                f"Agent {i}: {resp}" for i, resp in enumerate(agent_responses)
            )
            synthesis_prompt = "Synthesize these perspectives:\\n" + perspectives_text
            synthesis = provider.complete(synthesis_prompt, model=props.get("model"))
            shared[output_key] = {"synthesis": synthesis, "perspectives": agent_responses}

        elif node_type == "debate_advocate_node":
            if not provider:
                raise RuntimeError("Debate advocate requires a provider")
            position_key = str(props.get("position_key", "position"))
            topic_key = str(props.get("topic_key", "topic"))
            output_key = str(props.get("output_key", "argument"))
            position = str(props.get("position", "for"))
            topic = str(shared.get(topic_key, ""))
            prompt = "Argue " + position + " the following: " + topic + " Argument:"
            argument = provider.complete(prompt, model=props.get("model"))
            shared[output_key] = argument

        elif node_type == "debate_judge_node":
            if not provider:
                raise RuntimeError("Debate judge requires a provider")
            argument_a_key = str(props.get("argument_a_key", "argument_a"))
            argument_b_key = str(props.get("argument_b_key", "argument_b"))
            output_key = str(props.get("output_key", "verdict"))
            arg_a = str(shared.get(argument_a_key, ""))
            arg_b = str(shared.get(argument_b_key, ""))
            judge_prompt = (
                "Compare these arguments and choose the stronger one:\\n\\n"
                "Argument A:\\n" + arg_a + "\\n\\nArgument B:\\n" + arg_b + "\\n\\nVerdict:"
            )
            verdict = provider.complete(judge_prompt, model=props.get("model"))
            shared[output_key] = verdict
            if "a" in verdict.lower():
                chosen_action = "argument_a"
            elif "b" in verdict.lower():
                chosen_action = "argument_b"
            else:
                chosen_action = "tie"

        elif node_type == "web_search_node":
            query_key = str(props.get("query_key", "query"))
            output_key = str(props.get("output_key", "results"))
            api_key = os.environ.get("SEARCH_API_KEY", "")
            query = str(shared.get(query_key, ""))
            if not api_key:
                raise RuntimeError("web_search_node requires SEARCH_API_KEY environment variable")
            if not query:
                raise RuntimeError("web_search_node requires a query")
            try:
                search_url = "https://api.search.brave.com/res/v1/web/search?q=" + urllib.parse.quote(query) + "&count=10"
                req = urllib.request.Request(search_url, headers={"Accept": "application/json", "X-Subscription-Token": api_key})
                with urllib.request.urlopen(req) as resp:
                    results = json.loads(resp.read())
                shared[output_key] = results.get("web", [])
                chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "web_scrape_node":
            url_key = str(props.get("url_key", "url"))
            output_key = str(props.get("output_key", "content"))
            url = str(shared.get(url_key, ""))
            if not url:
                raise RuntimeError("web_scrape_node requires a URL")
            try:
                from bs4 import BeautifulSoup
                with urllib.request.urlopen(url) as resp:
                    html = resp.read().decode("utf-8")
                soup = BeautifulSoup(html, "html.parser")
                text = soup.get_text(separator=" ", strip=True)
                shared[output_key] = text[:5000]
                chosen_action = "success"
            except ImportError:
                raise RuntimeError("web_scrape_node requires beautifulsoup4: pip install beautifulsoup4")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "webhook_trigger_node":
            payload_key = str(props.get("payload_key", "payload"))
            output_key = str(props.get("output_key", "data"))
            webhook_path = str(props.get("webhook_path", "/webhook"))
            payload = shared.get(payload_key, {})
            if isinstance(payload, str):
                try:
                    payload = json.loads(payload)
                except json.JSONDecodeError:
                    payload = {"raw": payload}
            shared[output_key] = payload
            chosen_action = "received"

        elif node_type == "notification_node":
            message_key = str(props.get("message_key", "message"))
            channel = str(props.get("channel", "slack"))
            message = str(shared.get(message_key, ""))
            webhook_url = ""
            if channel == "slack":
                webhook_url = os.environ.get("SLACK_WEBHOOK", "")
            elif channel == "discord":
                webhook_url = os.environ.get("DISCORD_WEBHOOK", "")
            elif channel == "teams":
                webhook_url = os.environ.get("TEAMS_WEBHOOK", "")
            if not webhook_url:
                raise RuntimeError("notification_node requires webhook URL for " + channel)
            try:
                payload = json.dumps({"text": message}).encode()
                req = urllib.request.Request(webhook_url, data=payload, method="POST", headers={"Content-Type": "application/json"})
                with urllib.request.urlopen(req) as resp:
                    resp.read()
                chosen_action = "sent"
            except Exception as e:
                raise RuntimeError("Failed to send notification: " + str(e))

        elif node_type == "pdf_extract_node":
            file_path_key = str(props.get("file_path_key", "file_path"))
            output_key = str(props.get("output_key", "text"))
            file_path = str(shared.get(file_path_key, ""))
            if not file_path:
                raise RuntimeError("pdf_extract_node requires a file path")
            try:
                from PyPDF2 import PdfReader
                reader = PdfReader(file_path)
                text_parts = []
                for page in reader.pages:
                    text_parts.append(page.extract_text())
                shared[output_key] = " ".join(text_parts)
                chosen_action = "success"
            except ImportError:
                raise RuntimeError("pdf_extract_node requires PyPDF2: pip install PyPDF2")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "spreadsheet_node":
            file_path_key = str(props.get("file_path_key", "file_path"))
            operation = str(props.get("operation", "read"))
            sheet_name = str(props.get("sheet_name", "Sheet1"))
            output_key = str(props.get("output_key", "data"))
            file_path = str(shared.get(file_path_key, ""))
            if not file_path:
                raise RuntimeError("spreadsheet_node requires a file path")
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path)
                ws = wb[sheet_name]
                if operation == "read":
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        rows.append(list(row))
                    shared[output_key] = rows
                    chosen_action = "success"
                else:
                    raise RuntimeError("spreadsheet_node write operation not supported in standalone mode")
            except ImportError:
                raise RuntimeError("spreadsheet_node requires openpyxl: pip install openpyxl")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "secret_node":
            key = str(props.get("key", ""))
            output_key = str(props.get("output_key", "secret"))
            source = str(props.get("source", "env"))
            if not key:
                raise RuntimeError("secret_node requires a key")
            try:
                secret_value = ""
                if source == "env":
                    secret_value = os.environ.get(key, "")
                elif source == "dotenv":
                    try:
                        from dotenv import load_dotenv
                        load_dotenv()
                        secret_value = os.environ.get(key, "")
                    except ImportError:
                        raise RuntimeError("secret_node dotenv source requires python-dotenv: pip install python-dotenv")
                if secret_value:
                    shared[output_key] = secret_value
                    chosen_action = "found"
                else:
                    shared[f"{output_key}_error"] = "Secret not found: " + key
                    chosen_action = "not_found"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "registry_node":
            operation = str(props.get("operation", "set"))
            key = str(props.get("key", ""))
            input_key = str(props.get("input_key", "value"))
            output_key = str(props.get("output_key", "result"))
            registry_key = "__registry__"
            if registry_key not in shared:
                shared[registry_key] = {}
            registry = shared[registry_key]
            try:
                if operation == "set":
                    value = shared.get(input_key)
                    registry[key] = value
                    shared[output_key] = value
                    chosen_action = "set"
                elif operation == "get":
                    if key in registry:
                        value = registry[key]
                        shared[output_key] = value
                        chosen_action = "found"
                    else:
                        shared[f"{output_key}_error"] = "Key not in registry: " + key
                        chosen_action = "not_found"
                elif operation == "delete":
                    if key in registry:
                        del registry[key]
                        chosen_action = "deleted"
                    else:
                        chosen_action = "not_found"
                else:
                    chosen_action = "unknown_operation"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "stack_push_node":
            stack_name = str(props.get("stack_name", "default_stack"))
            input_key = str(props.get("input_key", "value"))
            value = shared.get(input_key)
            stack_key = "__stack_" + stack_name + "__"
            if stack_key not in shared:
                shared[stack_key] = []
            shared[stack_key].append(value)
            chosen_action = "pushed"

        elif node_type == "stack_pop_node":
            stack_name = str(props.get("stack_name", "default_stack"))
            output_key = str(props.get("output_key", "value"))
            stack_key = "__stack_" + stack_name + "__"
            if stack_key not in shared:
                shared[stack_key] = []
            stack = shared[stack_key]
            try:
                if stack:
                    value = stack.pop()
                    shared[output_key] = value
                    chosen_action = "popped"
                else:
                    shared[f"{output_key}_error"] = "Stack is empty: " + stack_name
                    chosen_action = "empty"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "queue_enqueue_node":
            queue_name = str(props.get("queue_name", "default_queue"))
            input_key = str(props.get("input_key", "value"))
            value = shared.get(input_key)
            queue_key = "__queue_" + queue_name + "__"
            if queue_key not in shared:
                shared[queue_key] = []
            shared[queue_key].append(value)
            chosen_action = "enqueued"

        elif node_type == "queue_dequeue_node":
            queue_name = str(props.get("queue_name", "default_queue"))
            output_key = str(props.get("output_key", "value"))
            queue_key = "__queue_" + queue_name + "__"
            if queue_key not in shared:
                shared[queue_key] = []
            queue = shared[queue_key]
            try:
                if queue:
                    value = queue.pop(0)
                    shared[output_key] = value
                    chosen_action = "dequeued"
                else:
                    shared[f"{output_key}_error"] = "Queue is empty: " + queue_name
                    chosen_action = "empty"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "local_memory_node":
            operation = str(props.get("operation", "set"))
            key = str(props.get("key", ""))
            input_key = str(props.get("input_key", "value"))
            output_key = str(props.get("output_key", "result"))
            memory_key = "__local_memory__"
            if memory_key not in shared:
                shared[memory_key] = {}
            memory = shared[memory_key]
            try:
                if operation == "set":
                    value = shared.get(input_key)
                    memory[key] = value
                    shared[output_key] = value
                    chosen_action = "set"
                elif operation == "get":
                    if key in memory:
                        value = memory[key]
                        shared[output_key] = value
                        chosen_action = "found"
                    else:
                        shared[f"{output_key}_error"] = "Key not in local memory: " + key
                        chosen_action = "not_found"
                elif operation == "clear":
                    memory.clear()
                    chosen_action = "cleared"
                else:
                    chosen_action = "unknown_operation"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "db_schema_node":
            db_type = str(props.get("db_type", "sqlite"))
            db_path = str(props.get("db_path", ""))
            output_key = str(props.get("output_key", "schema"))
            if not db_path:
                raise RuntimeError("db_schema_node requires a db_path")
            try:
                import sqlite3
                if db_type == "sqlite":
                    conn = sqlite3.connect(db_path)
                    cursor = conn.cursor()
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
                    tables = cursor.fetchall()
                    schema = {}
                    for table in tables:
                        table_name = table[0]
                        cursor.execute("PRAGMA table_info(" + table_name + ");")
                        columns = cursor.fetchall()
                        schema[table_name] = [{"name": col[1], "type": col[2]} for col in columns]
                    conn.close()
                    shared[output_key] = schema
                    chosen_action = "success"
                else:
                    raise RuntimeError("db_schema_node only supports sqlite in standalone mode")
            except ImportError:
                raise RuntimeError("db_schema_node requires sqlite3 (built-in)")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "nl_to_sql_node":
            if not provider:
                raise RuntimeError("nl_to_sql_node requires a provider")
            query_key = str(props.get("query_key", "query"))
            schema_key = str(props.get("schema_key", "schema"))
            output_key = str(props.get("output_key", "sql"))
            query = str(shared.get(query_key, ""))
            schema = shared.get(schema_key, {})
            if not query:
                raise RuntimeError("nl_to_sql_node requires a natural language query")
            try:
                schema_str = json.dumps(schema)
                prompt = (
                    "Convert this natural language query to SQL based on the database schema:"
                    "\\n\\nSchema:\\n" + schema_str + "\\n\\nQuery: " + query + "\\n\\nSQL:"
                )
                sql = provider.complete(prompt, model=props.get("model"))
                shared[output_key] = sql
                chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "sql_execute_node":
            db_type = str(props.get("db_type", "sqlite"))
            db_path = str(props.get("db_path", ""))
            sql_key = str(props.get("sql_key", "sql"))
            output_key = str(props.get("output_key", "results"))
            sql = str(shared.get(sql_key, ""))
            if not db_path:
                raise RuntimeError("sql_execute_node requires a db_path")
            if not sql:
                raise RuntimeError("sql_execute_node requires a sql query")
            try:
                import sqlite3
                if db_type == "sqlite":
                    conn = sqlite3.connect(db_path)
                    cursor = conn.cursor()
                    is_select = sql.strip().upper().startswith("SELECT")
                    cursor.execute(sql)
                    if is_select:
                        columns = [description[0] for description in cursor.description]
                        rows = cursor.fetchall()
                        results = [dict(zip(columns, row)) for row in rows]
                        shared[output_key] = results
                        chosen_action = "success"
                    else:
                        conn.commit()
                        shared[output_key] = {"affected_rows": cursor.rowcount}
                        chosen_action = "executed"
                    conn.close()
                else:
                    raise RuntimeError("sql_execute_node only supports sqlite in standalone mode")
            except ImportError:
                raise RuntimeError("sql_execute_node requires sqlite3 (built-in)")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "text_chunk_node":
            input_key = str(props.get("input_key", "text"))
            output_key = str(props.get("output_key", "chunks"))
            chunk_size = int(props.get("chunk_size", 1000))
            overlap = int(props.get("overlap", 0))
            text = str(shared.get(input_key, ""))
            if not text:
                raise RuntimeError("text_chunk_node requires input text")
            try:
                chunks = []
                i = 0
                while i < len(text):
                    chunk = text[i:i + chunk_size]
                    chunks.append(chunk)
                    i += chunk_size - overlap
                shared[output_key] = chunks
                chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "data_validate_node":
            input_key = str(props.get("input_key", "data"))
            output_key = str(props.get("output_key", "valid"))
            validation_type = str(props.get("validation_type", "type"))
            expected_type = str(props.get("expected_type", "str"))
            data = shared.get(input_key)
            try:
                is_valid = False
                if validation_type == "type":
                    if expected_type == "str":
                        is_valid = isinstance(data, str)
                    elif expected_type == "dict":
                        is_valid = isinstance(data, dict)
                    elif expected_type == "list":
                        is_valid = isinstance(data, list)
                    elif expected_type == "int":
                        is_valid = isinstance(data, int)
                    elif expected_type == "float":
                        is_valid = isinstance(data, (int, float))
                    else:
                        is_valid = True
                elif validation_type == "schema":
                    is_valid = isinstance(data, dict)
                else:
                    is_valid = data is not None
                shared[output_key] = is_valid
                chosen_action = "valid" if is_valid else "invalid"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "regex_node":
            input_key = str(props.get("input_key", "text"))
            output_key = str(props.get("output_key", "matches"))
            pattern = str(props.get("pattern", ""))
            operation = str(props.get("operation", "findall"))
            text = str(shared.get(input_key, ""))
            if not pattern:
                raise RuntimeError("regex_node requires a pattern")
            try:
                if operation == "findall":
                    matches = re.findall(pattern, text)
                    shared[output_key] = matches
                    chosen_action = "found" if matches else "not_found"
                elif operation == "sub":
                    replacement = str(props.get("replacement", ""))
                    result = re.sub(pattern, replacement, text)
                    shared[output_key] = result
                    chosen_action = "replaced"
                elif operation == "match":
                    match = re.match(pattern, text)
                    shared[output_key] = bool(match)
                    chosen_action = "matched" if match else "not_matched"
                else:
                    chosen_action = "unknown_operation"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "template_render_node":
            template_key = str(props.get("template_key", "template"))
            output_key = str(props.get("output_key", "rendered"))
            template = str(shared.get(template_key, ""))
            if not template:
                raise RuntimeError("template_render_node requires a template")
            try:
                result = template
                for key, value in shared.items():
                    if not key.startswith("_"):
                        result = result.replace("{{" + key + "}}", str(value))
                shared[output_key] = result
                chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "code_gen_node":
            if not provider:
                raise RuntimeError("code_gen_node requires a provider")
            spec_key = str(props.get("spec_key", "spec"))
            output_key = str(props.get("output_key", "code"))
            language = str(props.get("language", "python"))
            spec = str(shared.get(spec_key, ""))
            if not spec:
                raise RuntimeError("code_gen_node requires a specification")
            try:
                prompt = "Generate " + language + " code for the following specification:\\n\\n" + spec + "\\n\\nCode:"
                code = provider.complete(prompt, model=props.get("model"))
                shared[output_key] = code
                chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "python_tool_node":
            input_key = str(props.get("input_key", "input"))
            output_key = str(props.get("output_key", "output"))
            tool_path = str(props.get("tool_path", ""))
            if not tool_path:
                raise RuntimeError("python_tool_node requires a tool_path")
            try:
                import importlib.util
                spec = importlib.util.spec_from_file_location("tool", tool_path)
                if not spec or not spec.loader:
                    raise RuntimeError("Could not load tool from " + tool_path)
                tool = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(tool)
                if hasattr(tool, "run"):
                    result = tool.run(shared.get(input_key))
                    shared[output_key] = result
                    chosen_action = "success"
                else:
                    raise RuntimeError("Tool does not have a run() function")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "tty_serial_node":
            port = str(props.get("port", "/dev/ttyUSB0"))
            baudrate = int(props.get("baudrate", 9600))
            operation = str(props.get("operation", "read"))
            output_key = str(props.get("output_key", "data"))
            try:
                import serial
                ser = serial.Serial(port, baudrate, timeout=5)
                if operation == "read":
                    data = ser.readline().decode().strip()
                    shared[output_key] = data
                    chosen_action = "read" if data else "timeout"
                elif operation == "write":
                    input_key = str(props.get("input_key", "data"))
                    message = str(shared.get(input_key, ""))
                    ser.write(message.encode())
                    chosen_action = "written"
                ser.close()
            except ImportError:
                raise RuntimeError("tty_serial_node requires pyserial: pip install pyserial")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "test_gen_node":
            if not provider:
                raise RuntimeError("test_gen_node requires a provider")
            code_key = str(props.get("code_key", "code"))
            output_key = str(props.get("output_key", "tests"))
            test_framework = str(props.get("test_framework", "pytest"))
            code = str(shared.get(code_key, ""))
            if not code:
                raise RuntimeError("test_gen_node requires code to test")
            try:
                prompt = "Generate " + test_framework + " test cases for the following code:\\n\\n" + code + "\\n\\nTests:"
                tests = provider.complete(prompt, model=props.get("model"))
                shared[output_key] = tests
                chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "speech_to_text_node":
            audio_file_key = str(props.get("audio_file_key", "audio_file"))
            output_key = str(props.get("output_key", "text"))
            audio_file = str(shared.get(audio_file_key, ""))
            if not audio_file:
                raise RuntimeError("speech_to_text_node requires an audio file path")
            try:
                from speech_recognition import AudioFile, Recognizer
                recognizer = Recognizer()
                with AudioFile(audio_file) as source:
                    audio = recognizer.record(source)
                text = recognizer.recognize_google(audio)
                shared[output_key] = text
                chosen_action = "success"
            except ImportError:
                raise RuntimeError("speech_to_text_node requires SpeechRecognition: pip install SpeechRecognition")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "text_to_speech_node":
            input_key = str(props.get("input_key", "text"))
            output_key = str(props.get("output_key", "audio_file"))
            output_file = str(props.get("output_file", "output.mp3"))
            text = str(shared.get(input_key, ""))
            if not text:
                raise RuntimeError("text_to_speech_node requires text input")
            try:
                import pyttsx3
                engine = pyttsx3.init()
                engine.save_to_file(text, output_file)
                engine.runAndWait()
                shared[output_key] = output_file
                chosen_action = "success"
            except ImportError:
                raise RuntimeError("text_to_speech_node requires pyttsx3: pip install pyttsx3")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "image_vision_node":
            if not provider:
                raise RuntimeError("image_vision_node requires a provider")
            image_path_key = str(props.get("image_path_key", "image_path"))
            output_key = str(props.get("output_key", "description"))
            task = str(props.get("task", "describe"))
            image_path = str(shared.get(image_path_key, ""))
            if not image_path:
                raise RuntimeError("image_vision_node requires an image path")
            try:
                with open(image_path, "rb") as f:
                    image_data = f.read()
                import base64
                image_b64 = base64.b64encode(image_data).decode()
                prompt = "Analyze this image and " + task + ". Provide a detailed response."
                description = provider.complete(prompt, model=props.get("model"))
                shared[output_key] = description
                chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "embed_node":
            if not provider:
                raise RuntimeError("embed_node requires a provider")
            input_key = str(props.get("input_key", "text"))
            output_key = str(props.get("output_key", "embedding"))
            text = str(shared.get(input_key, ""))
            if not text:
                raise RuntimeError("embed_node requires text to embed")
            try:
                if hasattr(provider, "embed"):
                    embedding = provider.embed(text)
                    shared[output_key] = embedding
                else:
                    raise RuntimeError("Provider does not support embeddings")
                chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "vector_index_node":
            vectors_key = str(props.get("vectors_key", "vectors"))
            output_key = str(props.get("output_key", "index"))
            index_type = str(props.get("index_type", "simple"))
            vectors = shared.get(vectors_key, [])
            if not vectors:
                raise RuntimeError("vector_index_node requires vectors to index")
            try:
                index_data = {"type": index_type, "vectors": vectors, "ids": list(range(len(vectors)))}
                shared[output_key] = index_data
                chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "vector_retrieve_node":
            index_key = str(props.get("index_key", "index"))
            query_key = str(props.get("query_key", "query"))
            output_key = str(props.get("output_key", "results"))
            top_k = int(props.get("top_k", 5))
            index = shared.get(index_key, {})
            query = shared.get(query_key)
            if not index or not query:
                raise RuntimeError("vector_retrieve_node requires both index and query")
            try:
                vectors = index.get("vectors", [])
                if not vectors:
                    shared[output_key] = []
                    chosen_action = "not_found"
                else:
                    import math
                    def cosine_similarity(a, b):
                        if not a or not b:
                            return 0.0
                        dot_product = sum(x * y for x, y in zip(a, b))
                        mag_a = math.sqrt(sum(x * x for x in a))
                        mag_b = math.sqrt(sum(x * x for x in b))
                        if mag_a == 0 or mag_b == 0:
                            return 0.0
                        return dot_product / (mag_a * mag_b)

                    if isinstance(query, (list, tuple)):
                        query_vec = query
                    else:
                        query_vec = [query]

                    scores = []
                    for i, vec in enumerate(vectors):
                        sim = cosine_similarity(query_vec, vec)
                        scores.append({"id": i, "similarity": sim})
                    scores.sort(key=lambda x: x["similarity"], reverse=True)
                    results = scores[:top_k]
                    shared[output_key] = results
                    chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "retry_node":
            input_key = str(props.get("input_key", "input"))
            output_key = str(props.get("output_key", "result"))
            max_retries = int(props.get("max_retries", 3))
            backoff_factor = float(props.get("backoff_factor", 2.0))
            input_value = shared.get(input_key)
            try:
                shared[output_key] = input_value
                chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "rate_limiter_node":
            input_key = str(props.get("input_key", "input"))
            output_key = str(props.get("output_key", "result"))
            requests_per_second = float(props.get("requests_per_second", 1.0))
            input_value = shared.get(input_key)
            rate_limiter_key = "__rate_limiter__"
            if rate_limiter_key not in shared:
                shared[rate_limiter_key] = {"last_call": 0, "limit": requests_per_second}
            try:
                import time
                current_time = time.time()
                last_call = shared[rate_limiter_key]["last_call"]
                min_interval = 1.0 / requests_per_second
                if current_time - last_call < min_interval:
                    time.sleep(min_interval - (current_time - last_call))
                shared[rate_limiter_key]["last_call"] = time.time()
                shared[output_key] = input_value
                chosen_action = "allowed"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "email_send_node":
            recipient_key = str(props.get("recipient_key", "recipient"))
            subject_key = str(props.get("subject_key", "subject"))
            body_key = str(props.get("body_key", "body"))
            output_key = str(props.get("output_key", "status"))
            recipient = str(shared.get(recipient_key, ""))
            subject = str(shared.get(subject_key, ""))
            body = str(shared.get(body_key, ""))
            if not recipient or not subject or not body:
                raise RuntimeError("email_send_node requires recipient, subject, and body")
            try:
                import smtplib
                from email.mime.text import MIMEText
                from email.mime.multipart import MIMEMultipart
                sender = os.environ.get("EMAIL_ADDRESS", "")
                password = os.environ.get("EMAIL_PASSWORD", "")
                smtp_server = os.environ.get("SMTP_SERVER", "smtp.gmail.com")
                if not sender or not password:
                    raise RuntimeError("email_send_node requires EMAIL_ADDRESS and EMAIL_PASSWORD environment variables")
                msg = MIMEMultipart()
                msg["From"] = sender
                msg["To"] = recipient
                msg["Subject"] = subject
                msg.attach(MIMEText(body, "plain"))
                server = smtplib.SMTP(smtp_server, 587)
                server.starttls()
                server.login(sender, password)
                server.send_message(msg)
                server.quit()
                shared[output_key] = "sent"
                chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "email_read_node":
            output_key = str(props.get("output_key", "emails"))
            folder = str(props.get("folder", "INBOX"))
            max_emails = int(props.get("max_emails", 10))
            try:
                import imaplib
                from email.parser import Parser
                email_addr = os.environ.get("EMAIL_ADDRESS", "")
                password = os.environ.get("EMAIL_PASSWORD", "")
                imap_server = os.environ.get("IMAP_SERVER", "imap.gmail.com")
                if not email_addr or not password:
                    raise RuntimeError("email_read_node requires EMAIL_ADDRESS and EMAIL_PASSWORD environment variables")
                mail = imaplib.IMAP4_SSL(imap_server)
                mail.login(email_addr, password)
                mail.select(folder)
                status, messages = mail.search(None, "ALL")
                email_ids = messages[0].split()[-max_emails:]
                emails = []
                for email_id in email_ids:
                    status, msg_data = mail.fetch(email_id, "(RFC822)")
                    parser = Parser()
                    email_message = parser.parsestr(msg_data[0][1].decode())
                    emails.append({"subject": email_message.get("Subject"), "from": email_message.get("From"), "body": email_message.get_payload()})
                mail.close()
                mail.logout()
                shared[output_key] = emails
                chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "usb_serial_in_node":
            port = str(props.get("port", "/dev/ttyUSB0"))
            baudrate = int(props.get("baudrate", 9600))
            output_key = str(props.get("output_key", "data"))
            try:
                import serial
                ser = serial.Serial(port, baudrate, timeout=5)
                data = ser.readline().decode().strip()
                ser.close()
                shared[output_key] = data
                chosen_action = "success" if data else "timeout"
            except ImportError:
                raise RuntimeError("usb_serial_in_node requires pyserial: pip install pyserial")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "usb_serial_out_node":
            port = str(props.get("port", "/dev/ttyUSB0"))
            baudrate = int(props.get("baudrate", 9600))
            input_key = str(props.get("input_key", "data"))
            output_key = str(props.get("output_key", "status"))
            message = str(shared.get(input_key, ""))
            if not message:
                raise RuntimeError("usb_serial_out_node requires input data")
            try:
                import serial
                ser = serial.Serial(port, baudrate, timeout=5)
                ser.write(message.encode())
                ser.close()
                shared[output_key] = "sent"
                chosen_action = "success"
            except ImportError:
                raise RuntimeError("usb_serial_out_node requires pyserial: pip install pyserial")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "audio_input_node":
            output_key = str(props.get("output_key", "audio"))
            duration = float(props.get("duration", 5.0))
            sample_rate = int(props.get("sample_rate", 16000))
            try:
                import sounddevice
                import soundfile
                audio_data = sounddevice.rec(int(duration * sample_rate), samplerate=sample_rate, channels=1, dtype="float32")
                sounddevice.wait()
                output_file = str(props.get("output_file", "recording.wav"))
                soundfile.write(output_file, audio_data, sample_rate)
                shared[output_key] = output_file
                chosen_action = "success"
            except ImportError:
                raise RuntimeError("audio_input_node requires sounddevice and soundfile: pip install sounddevice soundfile")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "audio_output_node":
            input_key = str(props.get("input_key", "audio_file"))
            output_key = str(props.get("output_key", "status"))
            audio_file = str(shared.get(input_key, ""))
            if not audio_file:
                raise RuntimeError("audio_output_node requires an audio file path")
            try:
                import sounddevice
                import soundfile
                data, samplerate = soundfile.read(audio_file)
                sounddevice.play(data, samplerate)
                sounddevice.wait()
                shared[output_key] = "played"
                chosen_action = "success"
            except ImportError:
                raise RuntimeError("audio_output_node requires sounddevice and soundfile: pip install sounddevice soundfile")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "video_input_node":
            output_key = str(props.get("output_key", "video_file"))
            duration = float(props.get("duration", 5.0))
            output_file = str(props.get("output_file", "recording.mp4"))
            try:
                import cv2
                cap = cv2.VideoCapture(0)
                fourcc = cv2.VideoWriter_fourcc(*"mp4v")
                fps = cap.get(cv2.CAP_PROP_FPS)
                width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                out = cv2.VideoWriter(output_file, fourcc, fps, (width, height))
                frame_count = 0
                max_frames = int(duration * fps)
                while frame_count < max_frames:
                    ret, frame = cap.read()
                    if ret:
                        out.write(frame)
                        frame_count += 1
                    else:
                        break
                cap.release()
                out.release()
                shared[output_key] = output_file
                chosen_action = "success"
            except ImportError:
                raise RuntimeError("video_input_node requires opencv: pip install opencv-python")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "video_output_node":
            input_key = str(props.get("input_key", "video_file"))
            output_key = str(props.get("output_key", "status"))
            video_file = str(shared.get(input_key, ""))
            if not video_file:
                raise RuntimeError("video_output_node requires a video file path")
            try:
                import cv2
                cap = cv2.VideoCapture(video_file)
                while cap.isOpened():
                    ret, frame = cap.read()
                    if not ret:
                        break
                    cv2.imshow("Video", frame)
                    if cv2.waitKey(1) & 0xFF == ord("q"):
                        break
                cap.release()
                cv2.destroyAllWindows()
                shared[output_key] = "played"
                chosen_action = "success"
            except ImportError:
                raise RuntimeError("video_output_node requires opencv: pip install opencv-python")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "webcam_node":
            output_key = str(props.get("output_key", "image"))
            operation = str(props.get("operation", "capture"))
            try:
                import cv2
                cap = cv2.VideoCapture(0)
                if operation == "capture":
                    ret, frame = cap.read()
                    if ret:
                        output_file = str(props.get("output_file", "frame.jpg"))
                        cv2.imwrite(output_file, frame)
                        shared[output_key] = output_file
                        chosen_action = "success"
                    else:
                        raise RuntimeError("Failed to capture frame from webcam")
                elif operation == "stream":
                    frame_count = 0
                    max_frames = int(props.get("frame_count", 30))
                    frames = []
                    while frame_count < max_frames:
                        ret, frame = cap.read()
                        if ret:
                            frames.append(frame)
                            frame_count += 1
                        else:
                            break
                    shared[output_key] = frames
                    chosen_action = "success"
                else:
                    chosen_action = "unknown_operation"
                cap.release()
            except ImportError:
                raise RuntimeError("webcam_node requires opencv: pip install opencv-python")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "socket_node":
            operation = str(props.get("operation", "connect"))
            host = str(props.get("host", "localhost"))
            port = int(props.get("port", 5000))
            output_key = str(props.get("output_key", "result"))
            try:
                import socket as sock_module
                if operation == "connect":
                    input_key = str(props.get("input_key", "message"))
                    message = str(shared.get(input_key, ""))
                    s = sock_module.socket(sock_module.AF_INET, sock_module.SOCK_STREAM)
                    s.connect((host, port))
                    s.sendall(message.encode())
                    data = s.recv(4096).decode()
                    s.close()
                    shared[output_key] = data
                    chosen_action = "success"
                elif operation == "listen":
                    s = sock_module.socket(sock_module.AF_INET, sock_module.SOCK_STREAM)
                    s.setsockopt(sock_module.SOL_SOCKET, sock_module.SO_REUSEADDR, 1)
                    s.bind((host, port))
                    s.listen(1)
                    conn, addr = s.accept()
                    data = conn.recv(4096).decode()
                    conn.close()
                    s.close()
                    shared[output_key] = data
                    chosen_action = "received"
                else:
                    chosen_action = "unknown_operation"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "websocket_node":
            operation = str(props.get("operation", "connect"))
            url = str(props.get("url", "ws://localhost:8000"))
            output_key = str(props.get("output_key", "result"))
            try:
                import websockets
                if operation == "send":
                    input_key = str(props.get("input_key", "message"))
                    message = str(shared.get(input_key, ""))
                    import asyncio
                    async def send_msg():
                        async with websockets.connect(url) as ws:
                            await ws.send(message)
                            response = await ws.recv()
                            return response
                    loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(loop)
                    response = loop.run_until_complete(send_msg())
                    loop.close()
                    shared[output_key] = response
                    chosen_action = "success"
                else:
                    chosen_action = "unknown_operation"
            except ImportError:
                raise RuntimeError("websocket_node requires websockets: pip install websockets")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "a2a_send_node":
            recipient_key = str(props.get("recipient_key", "recipient"))
            message_key = str(props.get("message_key", "message"))
            output_key = str(props.get("output_key", "status"))
            recipient = str(shared.get(recipient_key, ""))
            message = str(shared.get(message_key, ""))
            if not recipient or not message:
                raise RuntimeError("a2a_send_node requires recipient and message")
            try:
                a2a_key = "__a2a_messages__"
                if a2a_key not in shared:
                    shared[a2a_key] = {}
                if recipient not in shared[a2a_key]:
                    shared[a2a_key][recipient] = []
                shared[a2a_key][recipient].append(message)
                shared[output_key] = "sent"
                chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "a2a_receive_node":
            sender_key = str(props.get("sender_key", "sender"))
            output_key = str(props.get("output_key", "message"))
            sender = str(shared.get(sender_key, ""))
            if not sender:
                raise RuntimeError("a2a_receive_node requires a sender identifier")
            try:
                a2a_key = "__a2a_messages__"
                if a2a_key not in shared:
                    shared[a2a_key] = {}
                if sender in shared[a2a_key] and shared[a2a_key][sender]:
                    message = shared[a2a_key][sender].pop(0)
                    shared[output_key] = message
                    chosen_action = "received"
                else:
                    shared[f"{output_key}_error"] = "No messages from " + sender
                    chosen_action = "empty"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "calendar_read_node":
            output_key = str(props.get("output_key", "events"))
            calendar_id = os.environ.get("GOOGLE_CALENDAR_ID", "")
            if not calendar_id:
                raise RuntimeError("calendar_read_node requires GOOGLE_CALENDAR_ID environment variable")
            try:
                from google.auth.transport.requests import Request
                from google.oauth2.credentials import Credentials
                from google_auth_oauthlib.flow import InstalledAppFlow
                from googleapiclient.discovery import build
                SCOPES = ["https://www.googleapis.com/auth/calendar.readonly"]
                creds = None
                if os.path.exists("token.json"):
                    creds = Credentials.from_authorized_user_file("token.json", SCOPES)
                if not creds or not creds.valid:
                    flow = InstalledAppFlow.from_client_secrets_file("credentials.json", SCOPES)
                    creds = flow.run_local_server(port=0)
                    with open("token.json", "w") as token:
                        token.write(creds.to_json())
                service = build("calendar", "v3", credentials=creds)
                events_result = service.events().list(calendarId=calendar_id, maxResults=10).execute()
                events = events_result.get("items", [])
                shared[output_key] = events
                chosen_action = "success"
            except ImportError:
                raise RuntimeError("calendar_read_node requires google-auth-oauthlib and google-api-python-client")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "calendar_write_node":
            calendar_id = os.environ.get("GOOGLE_CALENDAR_ID", "")
            event_key = str(props.get("event_key", "event"))
            output_key = str(props.get("output_key", "event_id"))
            event = shared.get(event_key, {})
            if not calendar_id or not event:
                raise RuntimeError("calendar_write_node requires GOOGLE_CALENDAR_ID and event data")
            try:
                from google.auth.transport.requests import Request
                from google.oauth2.credentials import Credentials
                from google_auth_oauthlib.flow import InstalledAppFlow
                from googleapiclient.discovery import build
                SCOPES = ["https://www.googleapis.com/auth/calendar"]
                creds = None
                if os.path.exists("token.json"):
                    creds = Credentials.from_authorized_user_file("token.json", SCOPES)
                if not creds or not creds.valid:
                    flow = InstalledAppFlow.from_client_secrets_file("credentials.json", SCOPES)
                    creds = flow.run_local_server(port=0)
                    with open("token.json", "w") as token:
                        token.write(creds.to_json())
                service = build("calendar", "v3", credentials=creds)
                result = service.events().insert(calendarId=calendar_id, body=event).execute()
                shared[output_key] = result.get("id")
                chosen_action = "success"
            except ImportError:
                raise RuntimeError("calendar_write_node requires google-auth-oauthlib and google-api-python-client")
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "mcp_tool_node":
            tool_name = str(props.get("tool_name", ""))
            input_key = str(props.get("input_key", "input"))
            output_key = str(props.get("output_key", "result"))
            server_url = os.environ.get("MCP_SERVER_URL", "http://localhost:3000")
            if not tool_name:
                raise RuntimeError("mcp_tool_node requires a tool_name")
            try:
                mcp_request = {
                    "jsonrpc": "2.0",
                    "id": 1,
                    "method": "tools/call",
                    "params": {"name": tool_name, "arguments": shared.get(input_key, {})},
                }
                data = json.dumps(mcp_request).encode()
                req = urllib.request.Request(
                    server_url + "/mcp",
                    data=data,
                    method="POST",
                    headers={"Content-Type": "application/json"},
                )
                with urllib.request.urlopen(req) as resp:
                    result = json.loads(resp.read())
                shared[output_key] = result.get("result", {})
                chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "shell_command_node":
            command = str(props.get("command", ""))
            output_key = str(props.get("output_key", "output"))
            shell_type = str(props.get("shell_type", "bash"))
            if not command:
                raise RuntimeError("shell_command_node requires a command")
            try:
                import subprocess
                result = subprocess.run(command, shell=True, capture_output=True, text=True, timeout=30)
                shared[output_key] = result.stdout
                if result.returncode == 0:
                    chosen_action = "success"
                else:
                    shared[f"{output_key}_error"] = result.stderr
                    chosen_action = "error"
            except subprocess.TimeoutExpired:
                shared[f"{output_key}_error"] = "Command timed out"
                chosen_action = "timeout"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "batch_node":
            items_key = str(props.get("items_key", "items"))
            operation_key = str(props.get("operation_key", "operation"))
            output_key = str(props.get("output_key", "results"))
            items = shared.get(items_key, [])
            if not isinstance(items, list):
                items = [items]
            try:
                results = []
                for item in items:
                    results.append(item)
                shared[output_key] = results
                chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "async_node":
            operation_key = str(props.get("operation_key", "operation"))
            input_key = str(props.get("input_key", "input"))
            output_key = str(props.get("output_key", "result"))
            input_value = shared.get(input_key)
            try:
                shared[output_key] = input_value
                chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "async_batch_node":
            items_key = str(props.get("items_key", "items"))
            operation_key = str(props.get("operation_key", "operation"))
            output_key = str(props.get("output_key", "results"))
            items = shared.get(items_key, [])
            if not isinstance(items, list):
                items = [items]
            try:
                results = []
                for item in items:
                    results.append(item)
                shared[output_key] = results
                chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "async_parallel_batch_node":
            items_key = str(props.get("items_key", "items"))
            operation_key = str(props.get("operation_key", "operation"))
            output_key = str(props.get("output_key", "results"))
            items = shared.get(items_key, [])
            if not isinstance(items, list):
                items = [items]
            try:
                results = []
                for item in items:
                    results.append(item)
                shared[output_key] = results
                chosen_action = "success"
            except Exception as e:
                shared[f"{output_key}_error"] = str(e)
                chosen_action = "error"

        elif node_type == "provider_failover_node":
            import json as _json
            prompt_key = str(props.get("prompt_key", "prompt"))
            output_key = str(props.get("output_key", "failover_response"))
            error_key = str(props.get("error_key", "failover_error"))
            config_str = str(props.get("providers_config", "[]"))
            try:
                config_list = _json.loads(config_str)
            except Exception:
                config_list = []

            if not config_list:
                shared[error_key] = "No providers configured"
                chosen_action = "all_failed"
            else:
                # Build FailoverProvider entries from config and available providers
                entries = []
                for cfg in config_list:
                    priority = int(cfg.get("priority", 999))
                    profile_id = cfg.get("profile_id", "")
                    provider_var = f"_provider_{profile_id.lower().replace('-', '_')}"

                    # Try to get the provider variable from locals (created by _render_provider_instances)
                    if provider_var in locals():
                        provider = locals()[provider_var]
                    elif profile_id in _graph_providers:
                        provider = _graph_providers[profile_id]
                    else:
                        continue

                    entries.append({
                        "priority": priority,
                        "provider": provider,
                        "model": cfg.get("model") or None,
                        "timeout_retries": int(cfg.get("timeout_retries", 3)),
                        "network_retries": int(cfg.get("network_retries", 3)),
                        "ratelimit_retries": int(cfg.get("ratelimit_retries", 2)),
                        "expired_retries": int(cfg.get("expired_retries", 1)),
                        "unknown_retries": int(cfg.get("unknown_retries", 1)),
                        "retry_delay": float(cfg.get("retry_delay", 2.0)),
                    })

                if entries:
                    cooldown_key = f"_pf_cooldown_{node['id']}"
                    if cooldown_key not in shared:
                        shared[cooldown_key] = {}
                    cooldowns = shared[cooldown_key]

                    failover = FailoverProvider(entries, cooldowns)
                    prompt = str(shared.get(prompt_key, ""))
                    try:
                        response = failover.complete(prompt)
                        shared[output_key] = response
                        chosen_action = "success"
                    except Exception as exc:
                        shared[error_key] = str(exc)
                        chosen_action = "all_failed"
                else:
                    shared[error_key] = "No valid providers resolved"
                    chosen_action = "all_failed"

        elif node_type == "trace_node":
            # Observability: emit OpenTelemetry span (no-op in standalone)
            # In standalone mode, tracing is logged to stderr for visibility
            span_name = str(props.get("span_name", node_id))
            keys_to_trace = str(props.get("keys_to_trace", "")).split(",")
            trace_data = {k.strip(): shared.get(k.strip()) for k in keys_to_trace if k.strip()}
            print(f"[TRACE] {span_name}: {trace_data}", file=sys.stderr)

        else:
            # Passthrough for unknown types
            pass

    except Exception as e:
        print(f"ERROR in {node['title']}: {e}", file=sys.stderr)
        raise

    return chosen_action'''

    def _render_run_flow(self) -> str:
        """Render the main run_flow() function."""
        return '''\
# ── Flow Runner ───────────────────────────────────────────────────────────────

def run_flow(shared=None):
    """Execute the flow and return the final shared store state."""
    shared = dict(shared or {})
    edge_map = {}
    for e in _EDGES:
        if e["from"] not in edge_map:
            edge_map[e["from"]] = []
        edge_map[e["from"]].append(e)

    current_id = _START
    visited = 0
    max_steps = 200

    while current_id and visited < max_steps:
        node = _NODES.get(current_id)
        if not node:
            break
        visited += 1

        outgoing = edge_map.get(current_id, [])
        outgoing_actions = {e["action"] for e in outgoing}

        try:
            chosen = _run_node(current_id, node, shared, outgoing_actions)
        except Exception as e:
            print(f"FATAL: Node '{node['title']}' failed: {e}", file=sys.stderr)
            raise

        # Find next edge by chosen action
        next_edge = next((e for e in outgoing if e["action"] == chosen), None)
        if next_edge is None and outgoing:
            # Fallback: try "default" edge, then first edge
            next_edge = next((e for e in outgoing if e["action"] == "default"), outgoing[0])

        current_id = next_edge["to"] if next_edge else ""

    return shared'''

    def _render_main_block(self, graph: GraphModel) -> str:
        """Render the __main__ block."""
        return f"""\
# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"Running: {graph.title}")
    try:
        result = run_flow({{}})
        print(f"\\n✓ Flow completed successfully")
        print("\\nFinal shared store:")
        print(json.dumps(result, indent=2, default=str))
    except Exception as e:
        print(f"\\n✗ Flow failed: {{e}}", file=sys.stderr)
        sys.exit(1)"""
